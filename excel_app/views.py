from django.shortcuts import render
from django.http import HttpResponse
import csv
from datetime import datetime
from .models import Product

# note == pip install openpyxl
import openpyxl
# from openpyxl import load_workbook

# import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook


# Create your views here.


def index(request):
    return HttpResponse('welcome to first page')


##### product MASTER EXCEL FILE UPLOAD ##########

def product_Upload(request):       
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES.get('excel_file', None)
        print(excel_file)
        if excel_file is not None:
            # Process the file
            print('hhhhhh')
            pass
        else:
            return HttpResponse('excel_file key is missing in the request.')
            
        excel_file_path = 'excel_file.xlsx'
        workbook = openpyxl.load_workbook(excel_file, data_only = True)
        
        # workbook = Workbook()
        # sheet = workbook.active
        
        if 'UploadFormat' in workbook.sheetnames:
            sheet = workbook['UploadFormat']
            print('sheet',sheet)
        else:
            print("Worksheet 'UploadFormat' does not exist in the workbook.")
        
        # sheet = workbook['UploadFormat']
        # print('sheet',sheet)
      
# field are used key and value mapping model field and excel field
        header_mapping = {
        #   Model field  and excel field
            'group':'group',
            'customer_id' :  'customer_id',
            'product_name' :  'product_name',
            'product_feature' :  'product_feature',
            
            'rate_basis' :  'rate_basis',
            'multiplying_factor' :  'multiplying_factor',
            'rate' :  'rate',
            'len_from' :  'len_from',
            'len_upto' :  'len_upto',
            'w_from' :  'w_from',
            'w_upto' :  'w_upto',
            'th_from' :  'th_from',
            'th_upto' :  'th_upto',

            'valid_from' :  'valid_from',
            'valid_upto' :  'valid_upto',
            'valid_yn' :  'valid_yn',
        }

# save data in list
        data_to_save =[]
        for row in sheet.iter_rows(min_row=2, values_only = True):
            print('row-----',row)
            data = {}
            for col, header  in enumerate(sheet[1]):
                field_name = header_mapping.get(header.value,None)
                if field_name:
                    data[field_name]= row[col]

            data_to_save.append(data)
        
        this_group = data_to_save[0]['group']
        print('this_group-  -->  ', this_group)

        this_valid_from =  data_to_save[0]['valid_from']
        past_valid_upto =  '2099-12-31'         # this_valid_from - timedelta(days=1)
        print('timedelta(days=1)-----', timedelta(days=1))
        print('this_valid_from-------', this_valid_from, past_valid_upto)


        rate_master_obj = Product.objects.filter(group = this_group, valid_yn = 'True')
        for rmo in  rate_master_obj:
            rmo.valid_yn = 'False'
            rmo.valid_upto = '2099-12-31'     #this_valid_from - timedelta(days=1)  #past_valid_upto  
            rmo.save()

        for row in data_to_save:
            print(row)
            rm_row = Product_Master()
            rm_row.group = row['group']
            rm_row.cust_code = row['customer_id']
            rm_row.product_name = row['product_name'].upper()
            rm_row.product_feature = row['product_feature'].upper().capitalize()
                    
            rm_row.rate_basis = row['rate_basis']
            rm_row.multiplying_factor = row['multiplying_factor']
            rm_row.rate = row['rate'] 
            rm_row.len_from = row['len_from']
            rm_row.len_upto = row['len_upto']
            rm_row.w_from = row['w_from']
            rm_row.w_upto = row['w_upto']
            rm_row.th_from = row['th_from']
            rm_row.th_upto = row['th_upto']
            
            rm_row.valid_from = row['valid_from']  # datetime.now()
            rm_row.valid_upto = '2099-12-31'
            rm_row.valid_yn = 'True'
            rm_row.save()

    return render(request, 'product_Upload.html')

